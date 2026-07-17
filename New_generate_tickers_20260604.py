#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font

INSTRUMENTS = {
    "SBIN": {
        "step": 10,
        "description": "Options on State Bank of India",
        "underlying_id": "INE062A01020",
        "venue_id": "XNSE",
    },
    "ICICIBANK": {
        "step": 10,
        "description": "Options on ICICI Bank Limited",
        "underlying_id": "INE090A01021",
        "venue_id": "XNSE",
    },
    "RELIANCE": {
        "step": 20,
        "description": "Options on Reliance Industries Limited",
        "underlying_id": "INE002A01018",
        "venue_id": "XNSE",
    },
    "KOTAKBANK": {
        "step": 2.5,
        "description": "Options on Kotak Mahindra Bank Limited",
        "underlying_id": "INE237A01036",
        "venue_id": "XNSE",
    },
    "HDFCBANK": {
        "step": 5,
        "description": "Options on HDFC Bank Limited",
        "underlying_id": "INE040A01034",
        "venue_id": "XNSE",
    },
    "ASIANPAINT": {
        "step": 20,
        "description": "Options on Asian Paints Limited",
        "underlying_id": "INE021A01026",
        "venue_id": "XNSE",
    },
    "BHARTIARTL": {
        "step": 20,
        "description": "Options on Bharti Airtel Limited",
        "underlying_id": "INE397D01024",
        "venue_id": "XNSE",
    },
    "AXISBANK": {
        "step": 10,
        "description": "Options on Axis Bank Limited",
        "underlying_id": "INE238A01034",
        "venue_id": "XNSE",
    },
    "INDUSINDBK": {
        "step": 10,
        "description": "Options on IndusInd Bank Limited",
        "underlying_id": "INE095A01012",
        "venue_id": "XNSE",
    },
    "DIXON": {
        "step": 50,
        "description": "Options on Dixon Technologies (India) Limited",
        "underlying_id": "INE935N01020",
        "venue_id": "XNSE",
    },
    "BSE": {
        "step": 100,
        "description": "Options on BSE Limited",
        "underlying_id": "INE118H01025",
        "venue_id": "XNSE",
    },
    "COFORGE": {
        "step": 20,
        "description": "Options on Coforge Limited",
        "underlying_id": "INE591G01025",
        "venue_id": "XNSE",
    },
    "M&M": {
        "step": 10,
        "description": "Options on Mahindra & Mahindra Limited",
        "underlying_id": "INE101A01026",
        "venue_id": "XNSE",
    },
    "MAXHEALTH": {
        "step": 10,
        "description": "Options on Max Healthcare Institute Limited",
        "underlying_id": "INE027H01010",
        "venue_id": "XNSE",
    },
    "MAZDOCK": {
        "step": 20,
        "description": "Options on Mazagon Dock Shipbuilders Limited",
        "underlying_id": "INE249Z01020",
        "venue_id": "XNSE",
    },
    "TVSMOTOR": {
        "step": 20,
        "description": "Options on TVS Motor Company Limited",
        "underlying_id": "INE494B01023",
        "venue_id": "XNSE",
    },
    "SENSEX": {
        "step": 100,
        "description": "Options on SENSEX index",
        "underlying_id": "",  # пусто
        "venue_id": "XBOM",
    },
    "MCX": {
        "step": 100,
        "description": "Options on Multi Commodity Exchange of India Limited",
        "underlying_id": "INE745G01043",
        "venue_id": "XNSE",
    },
    "APOLLOHOSP": {
        "step": 50,
        "description": "Options on Apollo Hospitals Enterprise Limited",
        "underlying_id": "INE437A01024",
        "venue_id": "XNSE",
    },
    "GODFRYPHLP": {
        "step": 100,
        "description": "Options on Godfrey Phillips India Limited",
        "underlying_id": "INE260B01028",
        "venue_id": "XNSE",
    },
    "INFY": {
        "step": 5,
        "description": "Options on Infosys Limited",
        "underlying_id": "INE009A01021",
        "venue_id": "XNSE",
    },
    "LTM": {
        "step": 50,
        "description": "Options on LTM Limited",
        "underlying_id": "INE214T01019",
        "venue_id": "XNSE",
    },
    "ABB": {
        "step": 50,
        "description": "Options on ABB India Limited",
        "underlying_id": "INE117A01022",
        "venue_id": "XNSE",
    },
    "HDFCLIFE": {
        "step": 5,
        "description": "Options on HDFC Life Insurance Company Limited",
        "underlying_id": "INE795G01014",
        "venue_id": "XNSE",
    },
    "INDIGO": {
        "step": 50,
        "description": "Options on InterGlobe Aviation Limited",
        "underlying_id": "INE646L01027",
        "venue_id": "XNSE",
    },
    "INDIANB": {
        "step": 10,
        "description": "Options on Indian Bank",
        "underlying_id": "INE562A01011",
        "venue_id": "XNSE",
    },
    "NAUKRI": {
        "step": 10,
        "description": "Options on Info Edge (India) Limited",
        "underlying_id": "INE663F01024",
        "venue_id": "XNSE",
    },
    "PFC": {
        "step": 5,
        "description": "Options on Power Finance Corporation Limited",
        "underlying_id": "INE134E01011",
        "venue_id": "XNSE",
    },
    "TITAN": {
        "step": 50,
        "description": "Options on Titan Company Limited",
        "underlying_id": "INE280A01028",
        "venue_id": "XNSE",
    },
    "MANAPPURAM": {
        "step": 5,
        "description": "Options on Manappuram Finance Limited",
        "underlying_id": "INE522D01027",
        "venue_id": "XNSE",
    },
    "TRENT": {
        "step": 20,
        "description": "Options on Trent Limited",
        "underlying_id": "INE849A01020",
        "venue_id": "XNSE",
    },
}

COLUMNS = [
    "No",
    "Platform",
    "ProductType",
    "Symbol",
    "Description",
    "BaseCurrency",
    "QuoteCurrency",
    "SettlementCurrency",
    "PlatformMultiplier",
    "ContractSize",
    "ClearingObligation",
    "CryptoAssets",
    "AssetClass",
    "ContractType",
    "CFICode",
    "VenueID",
    "UniqueProductID",
    "ProductID",
    "UnderlyingIDType",
    "UnderlyingID",
    "UnderlyingIndex",
    "UnderlyingIndexName",
    "PriceType",
    "PricingIdentifier",
]

SIDES = ["CE", "PE"]  # всегда CE+PE


# ---------- utils ----------
def round_to_step(value: float, step: int) -> int:
    return int(round(value / step) * step)

def parse_int(prompt: str, min_value: int | None = None) -> int:
    while True:
        s = input(prompt).strip()
        try:
            v = int(s)
            if min_value is not None and v < min_value:
                print(f"Нужно число >= {min_value}")
                continue
            return v
        except ValueError:
            print("Введите целое число.")

def parse_float(prompt: str) -> float:
    while True:
        s = input(prompt).strip().replace(",", ".")
        try:
            return float(s)
        except ValueError:
            print("Введите число (можно с точкой).")

def ask_yes_no(prompt: str) -> bool:
    # True = yes, False = no
    while True:
        s = input(prompt).strip().lower()
        if s in {"y", "yes", "д", "да"}:
            return True
        if s in {"n", "no", "н", "нет"}:
            return False
        print("Ответьте y/n (да/нет).")

def normalize_stock_expiry(exp: str) -> str:
    exp = exp.strip().upper()
    if len(exp) != 5:
        raise ValueError("Экспирация для акций должна быть формата 26FEB (ровно 5 символов).")
    yy = exp[:2]
    mon = exp[2:]
    if not yy.isdigit():
        raise ValueError("Первые 2 символа должны быть годом (например 26).")
    if mon not in {"JAN","FEB","MAR","APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC"}:
        raise ValueError("Месяц должен быть JAN/FEB/.../DEC.")
    return exp

def normalize_sensex_expiry(exp: str) -> str:
    exp = exp.strip()
    if not exp.isdigit() or len(exp) != 5:
        raise ValueError("Экспирация SENSEX должна быть 5 цифр, например 26205.")
    return exp

def ask_instruments() -> list[str]:
    keys = list(INSTRUMENTS.keys())
    print("Какие тикеры генерируем? Варианты:")
    for i, k in enumerate(keys, 1):
        print(f"  {i}) {k}")
    print("Ввод: номера через запятую (например 1,4,6) или 'all'.")

    while True:
        s = input("> ").strip().lower()
        if s == "all":
            return keys
        parts = [p.strip() for p in s.split(",") if p.strip()]
        try:
            idxs = [int(p) for p in parts]
            chosen = []
            for idx in idxs:
                if idx < 1 or idx > len(keys):
                    raise ValueError
                chosen.append(keys[idx - 1])

            out, seen = [], set()
            for x in chosen:
                if x not in seen:
                    out.append(x)
                    seen.add(x)
            return out
        except ValueError:
            print("Неверный ввод. Пример: 1,2,6 или all.")


def make_rows(symbol_base: str, expiry: str, strikes: list[int], meta: dict) -> list[list]:
    """
    CFI:
      - для SENSEX: CE=OPEICS, PE=OPEIPS
      - для остальных: CE=OPATCS, PE=OPATPS
    """
    rows = []
    for strike in strikes:
        for side in SIDES:
            if symbol_base == "SENSEX":
                cfi = "OPEICS" if side == "CE" else "OPEIPS"
            else:
                cfi = "OPATCS" if side == "CE" else "OPATPS"

            symbol = f"{symbol_base}{expiry}{strike}{side}"

            rows.append([
                None,                   # No (поставим позже)
                "DIVYA",                # Platform
                "OOF",                  # ProductType
                symbol,                 # Symbol
                meta["description"],    # Description (без скобок)
                "INR",                  # BaseCurrency
                "INR",                  # QuoteCurrency
                "INR",                  # SettlementCurrency
                "",                     # PlatformMultiplier
                "",                     # ContractSize
                "UKWN",                 # ClearingObligation
                "FLSE",                 # CryptoAssets
                "EQUI",                 # AssetClass
                "OPTN",                 # ContractType
                cfi,                    # CFICode
                meta["venue_id"],       # VenueID
                "",                     # UniqueProductID
                "",                     # ProductID
                "I",                    # UnderlyingIDType
                meta["underlying_id"],  # UnderlyingID
                "",                     # UnderlyingIndex
                "",                     # UnderlyingIndexName
                "MONE - Monetary",      # PriceType
                "NA",                   # PricingIdentifier
            ])
    return rows

def autosize_columns(ws):
    max_width = 60
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        max_len = len(str(col_name))
        for cell in ws[get_column_letter(col_idx)]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)

def apply_table_style(ws):
    align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = align
            cell.border = border

    header_font = Font(bold=True)
    for c in range(1, max_col + 1):
        ws.cell(row=1, column=c).font = header_font


# ---------- prices persistence ----------
def prices_file_path() -> Path:
    return Path(__file__).resolve().parent / "point9_prices.json"

def load_prices() -> dict:
    p = prices_file_path()
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}

def save_prices(prices: dict) -> None:
    p = prices_file_path()
    p.write_text(json.dumps(prices, ensure_ascii=False, indent=2), encoding="utf-8")


def main():
    chosen = ask_instruments()

    stock_exp = None
    sensex_exp = None

    if any(k != "SENSEX" for k in chosen):
        while True:
            try:
                stock_exp = normalize_stock_expiry(input("Экспирация для АКЦИЙ (например 26FEB): "))
                break
            except ValueError as e:
                print(e)

    if "SENSEX" in chosen:
        while True:
            try:
                sensex_exp = normalize_sensex_expiry(input("Экспирация для SENSEX (например 26205): "))
                break
            except ValueError as e:
                print(e)

    n = parse_int("Сколько страйков вверх/вниз от цены? (0..): ", min_value=0)

    stored = load_prices()
    prices: dict[str, float] = {}

    have_any_stored = any(k in stored for k in chosen)
    use_stored = False
    if have_any_stored:
        print("\nНайдены сохранённые цены (point9_prices.json).")
        changed = ask_yes_no("Цены изменились? (y/n): ")
        use_stored = (not changed)

    for inst in chosen:
        meta = INSTRUMENTS[inst]
        step = meta["step"]
        expiry = sensex_exp if inst == "SENSEX" else stock_exp

        print("\n" + "=" * 60)
        print(f"[{inst}] step={step}, expiry={expiry}")

        if use_stored and inst in stored:
            center_price = float(stored[inst])
            print(f"[{inst}] Беру сохранённую цену: {center_price}")
        else:
            center_price = float(input(f"[{inst}] Цена актива (число): ").strip().replace(",", "."))

        prices[inst] = center_price

    stored_updated = dict(stored)
    stored_updated.update(prices)
    save_prices(stored_updated)

    all_rows = []
    for inst in chosen:
        meta = INSTRUMENTS[inst]
        step = meta["step"]
        expiry = sensex_exp if inst == "SENSEX" else stock_exp
        center_price = prices[inst]

        rounded_center = round_to_step(center_price, step)
        strikes = [rounded_center + i * step for i in range(-n, n + 1)]

        all_rows.extend(make_rows(inst, expiry, strikes, meta))

        print(f"[{inst}] center={center_price} -> rounded={rounded_center}")
        print(f"[{inst}] strikes: {strikes[0]} ... {strikes[-1]} (count={len(strikes)})")
        print(f"[{inst}] rows added: {len(strikes) * len(SIDES)}")

    all_rows.sort(key=lambda r: r[3])

    for i, r in enumerate(all_rows, start=1):
        r[0] = i

    wb = Workbook()
    ws = wb.active
    ws.title = "point9"

    ws.append(COLUMNS)
    for r in all_rows:
        ws.append(r)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

    autosize_columns(ws)
    apply_table_style(ws)

    out_name = f"point9_tickers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out_name)

    print("\nГотово.")
    print(f"Excel сохранён: {out_name}")
    print(f"Цены сохранены в: {prices_file_path().name}")


if __name__ == "__main__":
    main()
